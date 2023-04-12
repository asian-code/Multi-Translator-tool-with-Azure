import requests, uuid, json
import openpyxl
import re
from tkinter import Tk # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
# Add your key and endpoint
key = "4eb9af8481a140649991eebab08d3597"
endpoint = "https://api.cognitive.microsofttranslator.com/translate"
Tk ().withdraw ()
# location, also known as region.
# required if you're using a multi-service or regional (not global) resource. It can be found in the Azure portal on the Keys and Endpoint page.
location = "eastus"
cyan = '\033[36m'
reset = '\033[0m'

lang={
"Spanish" : "es",
"Portuguese" : "pt",
"French" : "fr",
"Slovak" : "sk",
"Chinese Simplified" : "zh-Hans" ,
"German" : "de",
"Italian" : "it",
"Hungarian" : "hu",
"Russian" : "ru",
"Polish" : "pl",
"Korean" : "ko",
"Japanese" : "ja",
"Hindi" : "hi",
"Indonesian" : "id",
"Chinese Traditional" : "zh-Hant",
"Dutch" : "nl",
"Norwegian" : "nb",
"Swedish" : "sv",
"Czech" : "cs",
"Danish" : "da"
}
lang_code=list(lang.values())
lang_name=list(lang.keys())
params = {
    'api-version': '3.0',
    'from': 'en',
    'to': list(lang.values())
}

headers = {
    'Ocp-Apim-Subscription-Key': key,
    # location required if you're using a multi-service or regional (not global) resource.
    'Ocp-Apim-Subscription-Region': location,
    'Content-type': 'application/json',
    'X-ClientTraceId': str(uuid.uuid4())
}
# read from the source file and store all IDs into list
sourcefile = askopenfilename ()
toTranslate=[]
print (sourcefile)
sourceSheet = openpyxl.load_workbook(sourcefile).active
for row in range(1, sourceSheet.max_row+1):
    for idx, cell in enumerate(sourceSheet[row]):
        if (cell.value is not None):  # only check if the cell is not empty
                if (cell.col_idx == 1):
                    toTranslate.append(cell.value)
pattern=r"#(\w+)"
print("\n\n",toTranslate,"\n\n")

#loop through all the text that need to be translated -----------------------------------------------------
#create results excel file
file=openpyxl.Workbook()
sheet=file.active
sheet.append(["English"]+lang_name)

def replace(match):
    placeholders.append(match.group(0))# save the placeholder
    return "[[]]"

for text in toTranslate:# PreTranslation processing
    
    placeholders=[]
    body = [{'text': re.sub(pattern, replace, text)}]
    #debuging
    print(placeholders)
    print(body[0])


    # get the translation of the text
    request = requests.post(endpoint, params=params, headers=headers, json=body)
    response = request.json()
    translations = response[0]["translations"]
    # print(json.dumps(response, sort_keys=True, ensure_ascii=False, indent=4, separators=(',', ': ')))
    
    row=[text]# adds the original text 
    for translation in translations:
        
        t=translation["text"]
        # PostTranslation processing
        for i in placeholders:
            t=t.replace("[[]]",i,1)
        row.append(t) 
        print("Done for "+lang_name[lang_code.index(translation["to"])])

    sheet.append(row)




file.save("translate_Results.xlsx")
print("Saved results to translate_Results.xlsx")

# print(json.dumps(response, sort_keys=True, ensure_ascii=False, indent=4, separators=(',', ': ')))